from __future__ import annotations

import argparse
import json
import re
import time
import unicodedata
from pathlib import Path
from typing import Any
from urllib.parse import quote, urljoin

from openpyxl import load_workbook
from requests import Response, Session
from requests.exceptions import RequestException

from data_pipeline_utils import (
    PROCESSED_DIR,
    RAW_DIR,
    build_retry_session,
    configure_utf8_stdio,
    content_hash,
    decode_response,
    ensure_data_dirs,
    normalize_unicode,
    write_json,
)


SITE_ROOT = "https://vbpl.vn"
DETAIL_URL = SITE_ROOT + "/van-ban/chi-tiet/x--{document_id}?tabs=tai-ve"
DOWNLOAD_ROOT = (
    "https://vbpl-bientap-gateway.moj.gov.vn/api/qtdc/public/doc/minio/"
    "buckets/vbpl/"
)
DEFAULT_EXPORT = RAW_DIR / "vbpl" / "vbpl_ho_tro_doanh_nghiep_trung_uong_con_hieu_luc.xlsx"
DEFAULT_INVENTORY = RAW_DIR / "vbpl" / "attachment_inventory.json"
DEFAULT_OUTPUT = PROCESSED_DIR / "vbpl_ho_tro_doanh_nghiep_manifest.json"
DEFAULT_FILES_DIR = RAW_DIR / "vbpl" / "files"
FILE_PARTS = [1, 2, 4]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Download every original/attachment file for the exact VBPL export: "
            "title contains 'hỗ trợ doanh nghiệp', central, active."
        )
    )
    parser.add_argument("--export", type=Path, default=DEFAULT_EXPORT)
    parser.add_argument("--inventory", type=Path, default=DEFAULT_INVENTORY)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--files-dir", type=Path, default=DEFAULT_FILES_DIR)
    parser.add_argument("--expected-count", type=int, default=47)
    parser.add_argument("--timeout", type=float, default=60.0)
    parser.add_argument("--retries", type=int, default=4)
    parser.add_argument("--delay", type=float, default=0.25)
    parser.add_argument(
        "--inventory-only",
        action="store_true",
        help="Collect attachment metadata without downloading the file bodies.",
    )
    return parser.parse_args()


def normalize_cell(value: Any) -> Any:
    if isinstance(value, str):
        return normalize_unicode(value.strip())
    return value


def load_official_export(path: Path, expected_count: int | None = 47) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [normalize_cell(value) for value in next(rows)]
    documents = [
        {str(headers[index]): normalize_cell(value) for index, value in enumerate(row)}
        for row in rows
        if any(value is not None for value in row)
    ]
    workbook.close()

    ids = [str(document.get("ID", "")).strip() for document in documents]
    if not all(ids) or len(ids) != len(set(ids)):
        raise ValueError("The official export contains blank or duplicate document IDs")
    if expected_count is not None and len(documents) != expected_count:
        raise ValueError(f"Expected {expected_count} documents, found {len(documents)}")

    invalid = [
        document
        for document in documents
        if document.get("Tình trạng hiệu lực") != "Còn hiệu lực"
    ]
    if invalid:
        raise ValueError(f"Found {len(invalid)} documents that are not active")
    return documents


def _script_urls(html: str, base_url: str) -> list[str]:
    sources = re.findall(r'<script[^>]+src="([^"]+)"', html, flags=re.IGNORECASE)
    return [urljoin(base_url, source) for source in sources]


def discover_files_action(session: Session, sample_document_id: str, timeout: float) -> str:
    detail_url = DETAIL_URL.format(document_id=sample_document_id)
    response = session.get(detail_url, timeout=timeout)
    response.raise_for_status()
    for script_url in _script_urls(response.text, detail_url):
        script_response = session.get(script_url, timeout=timeout)
        script_response.raise_for_status()
        script = script_response.text
        if '["document",t,"files","parts"' not in script:
            continue
        module_start = script.find("95743:function")
        module = script[module_start : module_start + 5000]
        action_match = re.search(r'\("([0-9a-f]{40})"\)', module)
        if action_match:
            return action_match.group(1)
    raise RuntimeError("Could not discover the current VBPL file-list action")


def parse_rsc_file_list(payload: str) -> list[dict[str, Any]]:
    for line in payload.splitlines():
        match = re.match(r"^\d+:(\[.*\])$", line)
        if not match:
            continue
        try:
            value = json.loads(match.group(1))
        except json.JSONDecodeError:
            continue
        if isinstance(value, list) and all(isinstance(item, dict) for item in value):
            if not value or any("fileName" in item for item in value):
                return value
    raise ValueError("VBPL returned an unexpected file-list response")


def fetch_file_list(
    session: Session,
    document_id: str,
    action_id: str,
    timeout: float,
    retries: int,
) -> list[dict[str, Any]]:
    detail_url = DETAIL_URL.format(document_id=document_id)
    for attempt in range(retries + 1):
        try:
            response = session.post(
                detail_url,
                headers={
                    "Accept": "text/x-component",
                    "Content-Type": "text/plain;charset=UTF-8",
                    "Next-Action": action_id,
                    "Referer": detail_url,
                },
                data=json.dumps([document_id, FILE_PARTS], ensure_ascii=False).encode("utf-8"),
                timeout=timeout,
            )
            response.raise_for_status()
            return parse_rsc_file_list(decode_response(response))
        except (RequestException, ValueError):
            if attempt >= retries:
                raise
            time.sleep(min(0.5 * (2**attempt), 8.0))
    raise RuntimeError("Unreachable retry state")


def safe_filename(value: str) -> str:
    name = unicodedata.normalize("NFC", value).replace("/", "_").replace("\\", "_")
    name = re.sub(r'[<>:"|?*\x00-\x1f]', "_", name).strip(" .")
    return name or "attachment.bin"


def stable_download_url(full_path: str) -> str:
    return DOWNLOAD_ROOT + quote(normalize_unicode(full_path), safe="/") + "/download"


def download_response(
    session: Session,
    url: str,
    fallback_url: str | None,
    timeout: float,
) -> Response:
    response = session.get(url, stream=True, timeout=timeout)
    if response.ok:
        return response
    response.close()
    if not fallback_url:
        response.raise_for_status()
    fallback = session.get(fallback_url, stream=True, timeout=timeout)
    fallback.raise_for_status()
    return fallback


def save_attachment(
    session: Session,
    attachment: dict[str, Any],
    document_id: str,
    files_dir: Path,
    timeout: float,
) -> tuple[Path, str, int, str]:
    filename = safe_filename(str(attachment.get("fileName") or "attachment.bin"))
    destination = files_dir / document_id / filename
    expected_size = int(attachment.get("size") or 0)
    full_path = str(attachment.get("fullPath") or f"{document_id}/{filename}")
    url = stable_download_url(full_path)

    if destination.exists() and (not expected_size or destination.stat().st_size == expected_size):
        payload = destination.read_bytes()
        return destination, content_hash(payload), len(payload), url

    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(destination.suffix + ".part")
    presigned_url = str(attachment.get("presignedUrl") or "") or None
    response = download_response(
        session,
        presigned_url or url,
        url if presigned_url else None,
        timeout,
    )
    try:
        with temporary.open("wb") as output:
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    output.write(chunk)
    finally:
        response.close()

    actual_size = temporary.stat().st_size
    if expected_size and actual_size != expected_size:
        temporary.unlink(missing_ok=True)
        raise ValueError(
            f"Size mismatch for {document_id}/{filename}: "
            f"expected {expected_size}, downloaded {actual_size}"
        )
    temporary.replace(destination)
    payload = destination.read_bytes()
    return destination, content_hash(payload), len(payload), url


def relative_to_project(path: Path) -> str:
    project_root = Path(__file__).resolve().parents[1]
    try:
        return path.resolve().relative_to(project_root).as_posix()
    except ValueError:
        return str(path.resolve())


def inventory_without_presigned_urls(
    inventory: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    return [
        {
            **document,
            "attachments": [
                {key: value for key, value in attachment.items() if key != "presignedUrl"}
                for attachment in document["attachments"]
            ],
        }
        for document in inventory
    ]


def collect_inventory(
    session: Session,
    documents: list[dict[str, Any]],
    action_id: str,
    timeout: float,
    retries: int,
    delay: float,
) -> list[dict[str, Any]]:
    inventory: list[dict[str, Any]] = []
    for index, source in enumerate(documents, start=1):
        document_id = str(source["ID"])
        attachments = fetch_file_list(session, document_id, action_id, timeout, retries)
        inventory.append(
            {
                "documentId": document_id,
                "detailUrl": DETAIL_URL.format(document_id=document_id),
                "source": source,
                "attachments": attachments,
            }
        )
        print(f"[{index:02d}/{len(documents)}] {document_id}: {len(attachments)} tệp")
        if delay:
            time.sleep(delay)
    return inventory


def download_inventory(
    session: Session,
    inventory: list[dict[str, Any]],
    files_dir: Path,
    timeout: float,
    delay: float,
) -> tuple[list[dict[str, Any]], int, int]:
    hashes: dict[str, str] = {}
    total_bytes = 0
    duplicate_count = 0
    attachment_count = sum(len(item["attachments"]) for item in inventory)
    position = 0

    for document in inventory:
        document_id = str(document["documentId"])
        processed_attachments: list[dict[str, Any]] = []
        for attachment in document["attachments"]:
            position += 1
            path, sha256, size, url = save_attachment(
                session, attachment, document_id, files_dir, timeout
            )
            local_path = relative_to_project(path)
            duplicate_of = hashes.get(sha256)
            if duplicate_of and duplicate_of != local_path:
                path.unlink(missing_ok=True)
                local_path = duplicate_of
                duplicate_count += 1
            else:
                hashes[sha256] = local_path
                total_bytes += size

            processed_attachments.append(
                {
                    "fileName": normalize_cell(attachment.get("fileName")),
                    "mediaType": attachment.get("type"),
                    "category": attachment.get("relatedType"),
                    "sourceSize": attachment.get("size"),
                    "downloadedSize": size,
                    "sha256": sha256,
                    "sourceUrl": url,
                    "localPath": local_path,
                    "duplicateOf": duplicate_of,
                }
            )
            print(f"  [{position:03d}/{attachment_count}] {document_id}/{attachment.get('fileName')}")
            if delay:
                time.sleep(delay)
        document["attachments"] = processed_attachments
    return inventory, total_bytes, duplicate_count


def main() -> None:
    args = parse_args()
    configure_utf8_stdio()
    ensure_data_dirs()
    if not args.export.exists():
        raise FileNotFoundError(
            f"Missing official VBPL export: {args.export}. "
            "Export the exact filtered result from vbpl.vn first."
        )

    documents = load_official_export(args.export, args.expected_count)
    session = build_retry_session(retries=args.retries)
    session.headers.update({"User-Agent": "GrantPilotAI-data-pipeline/1.0 (+https://vbpl.vn)"})
    action_id = discover_files_action(session, str(documents[0]["ID"]), args.timeout)
    inventory = collect_inventory(
        session, documents, action_id, args.timeout, args.retries, args.delay
    )
    raw_payload = {
        "source": SITE_ROOT,
        "query": {
            "keyword": "hỗ trợ doanh nghiệp",
            "searchIn": "Tiêu đề",
            "scope": "Trung ương",
            "status": "Còn hiệu lực",
        },
        "documentCount": len(inventory),
        "attachmentCount": sum(len(item["attachments"]) for item in inventory),
        "totalDeclaredBytes": sum(
            int(attachment.get("size") or 0)
            for item in inventory
            for attachment in item["attachments"]
        ),
        "documents": inventory_without_presigned_urls(inventory),
    }
    write_json(args.inventory, raw_payload)

    if args.inventory_only:
        print(f"Đã ghi inventory: {relative_to_project(args.inventory)}")
        return

    processed, total_bytes, duplicate_count = download_inventory(
        session, inventory, args.files_dir, args.timeout, args.delay
    )
    result = {
        "source": SITE_ROOT,
        "query": raw_payload["query"],
        "officialExport": relative_to_project(args.export),
        "documentCount": len(processed),
        "attachmentCount": sum(len(item["attachments"]) for item in processed),
        "uniqueFileBytes": total_bytes,
        "duplicateFileCount": duplicate_count,
        "documents": processed,
    }
    write_json(args.output, result)
    print(f"Đã ghi manifest: {relative_to_project(args.output)}")
    print(f"{len(processed)} văn bản, {result['attachmentCount']} tệp, {duplicate_count} tệp trùng")


if __name__ == "__main__":
    main()
